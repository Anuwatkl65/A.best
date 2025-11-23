from datetime import datetime
import openpyxl
from django.core.management.base import BaseCommand
from django.db import transaction

from production.models import Lot, ScanRecord, Machine


class Command(BaseCommand):
    help = "Import data from 'A.Best - Production Tracker.xlsx' into Django models"

    def add_arguments(self, parser):
        parser.add_argument(
            "path",
            nargs="?",
            default="A.Best - Production Tracker.xlsx",
            help="Path to the Excel file",
        )

    def handle(self, *args, **options):
        path = options["path"]
        self.stdout.write(self.style.NOTICE(f"Loading workbook: {path}"))
        wb = openpyxl.load_workbook(path, data_only=True)

        with transaction.atomic():
            self._import_machines(wb)
            lot_map = self._import_lots(wb)
            self._import_collect(wb, lot_map)

        self.stdout.write(self.style.SUCCESS("Import completed."))

    # ------------------------ Machines ------------------------

    def _import_machines(self, wb):
        if "Machine List" not in wb.sheetnames:
            self.stdout.write("Sheet 'Machine List' not found, skip machines.")
            return

        ws = wb["Machine List"]
        headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        idx = {name: i for i, name in enumerate(headers) if name}

        count = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            machine_no = row[idx.get("Machine No.")] if "Machine No." in idx else None
            if not machine_no:
                continue

            machine_name = row[idx.get("Machine Name")] if "Machine Name" in idx else ""
            department = row[idx.get("Department")] if "Department" in idx else ""

            Machine.objects.update_or_create(
                machine_no=str(machine_no).strip(),
                defaults={
                    "machine_name": str(machine_name or "").strip(),
                    "department": str(department or "").strip(),
                },
            )
            count += 1

        self.stdout.write(self.style.SUCCESS(f"Imported/updated {count} machines."))

    # ------------------------ Lots (Databased) ------------------------

    def _import_lots(self, wb):
        if "Databased" not in wb.sheetnames:
            self.stdout.write("Sheet 'Databased' not found, skip lots.")
            return {}

        ws = wb["Databased"]
        headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        idx = {name: i for i, name in enumerate(headers) if name}

        lot_map = {}
        count = 0

        for row in ws.iter_rows(min_row=2, values_only=True):
            lot_no = row[idx.get("Lot No.")] if "Lot No." in idx else None
            if not lot_no:
                continue

            abest_part_no = row[idx.get("A.Best Part No.")] if "A.Best Part No." in idx else ""
            customer = row[idx.get("Customer")] if "Customer" in idx else ""
            description = row[idx.get("Description")] if "Description" in idx else ""
            customer_part_no = row[idx.get("Customer Part No.")] if "Customer Part No." in idx else ""
            po_no = row[idx.get("PO No.")] if "PO No." in idx else ""
            remark = row[idx.get("Remark")] if "Remark" in idx else ""
            department = row[idx.get("Department")] if "Department" in idx else ""
            machine_no = row[idx.get("Machine No.")] if "Machine No." in idx else ""
            lot_type = row[idx.get("Type")] if "Type" in idx else "Order"

            prod_qty = row[idx.get("Production Quantity")] if "Production Quantity" in idx else 0
            pieces_per_box = row[idx.get("จำนวนบรรจุต่อกล่อง")] if "จำนวนบรรจุต่อกล่อง" in idx else 0
            target = row[idx.get("Target")] if "Target" in idx else None

            try:
                prod_qty = int(prod_qty or 0)
            except:
                prod_qty = 0
            try:
                pieces_per_box = int(pieces_per_box or 0)
            except:
                pieces_per_box = 0
            try:
                target = int(target) if target is not None else prod_qty
            except:
                target = prod_qty

            lot, created = Lot.objects.update_or_create(
                lot_no=str(lot_no).strip(),
                defaults={
                    "part_no": str(abest_part_no or "").strip(),
                    "customer": str(customer or "").strip(),
                    "description": str(description or "").strip(),
                    "customer_part_no": str(customer_part_no or "").strip(),
                    "po_no": str(po_no or "").strip(),
                    "remark": str(remark or "").strip(),
                    "production_quantity": prod_qty,
                    "pieces_per_box": pieces_per_box,
                    "target": target,
                    "department": str(department or "").strip(),
                    "machine_no": str(machine_no or "").strip(),
                    "type": str(lot_type or "Order").strip(),
                },
            )
            lot_map[lot.lot_no] = lot
            count += 1

        self.stdout.write(self.style.SUCCESS(f"Imported/updated {count} lots."))
        return lot_map

    # ------------------------ Scan Records (Collect) ------------------------

    def _import_collect(self, wb, lot_map):
        if "Collect" not in wb.sheetnames:
            self.stdout.write("Sheet 'Collect' not found, skip scan records.")
            return

        ws = wb["Collect"]
        headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        idx = {name: i for i, name in enumerate(headers) if name}

        count = 0
        first_scan = {}
        last_scan = {}

        for row in ws.iter_rows(min_row=2, values_only=True):
            lot_no = row[idx.get("Lot No.")] if "Lot No." in idx else None
            if not lot_no:
                continue

            lot_no = str(lot_no).strip()
            lot = lot_map.get(lot_no)
            if not lot:
                continue

            date_val = row[idx.get("Date")] if "Date" in idx else None
            time_val = row[idx.get("Time")] if "Time" in idx else None
            machine_no = row[idx.get("Machine No.")] if "Machine No." in idx else ""

            if time_val:
                scan_dt = datetime.combine(d, time_val)
            else:
                scan_dt = datetime.combine(d, datetime.min.time())

            # >>> เพิ่มตรงนี้ให้ datetime มี timezone <<<
            if timezone.is_naive(scan_dt):
                scan_dt = timezone.make_aware(scan_dt)

            qty = lot.pieces_per_box or 0   # 1 scan = 1 กล่อง

            ScanRecord.objects.create(
                lot=lot,
                machine_no=str(machine_no or "").strip(),
                qty=qty,
                scanned_at=scan_dt,
            )