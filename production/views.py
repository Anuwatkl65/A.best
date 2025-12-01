from datetime import datetime, timedelta

import openpyxl
import pandas as pd

from django.contrib import messages
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required, user_passes_test
from django.core.management.base import BaseCommand
from django.db import transaction
from django.db.models import Sum, Q
from django.db.models.functions import TruncDate, TruncHour, TruncMonth, Coalesce
from django.http import JsonResponse, HttpResponse
from django.shortcuts import render, redirect, get_object_or_404
from django.urls import reverse
from django.utils import timezone
from django.views.decorators.csrf import csrf_exempt
from django.contrib.sessions.models import Session
from django.contrib.auth.models import User
from django.utils.timezone import now


from openpyxl.utils import get_column_letter

from .models import Lot, ScanRecord, UserProfile, Machine


MONTH_TH = {
    1: "ม.ค.", 2: "ก.พ.", 3: "มี.ค.", 4: "เม.ย.",
    5: "พ.ค.", 6: "มิ.ย.", 7: "ก.ค.", 8: "ส.ค.",
    9: "ก.ย.", 10: "ต.ค.", 11: "พ.ย.", 12: "ธ.ค.",
}

# label ชื่อแผนก
LABELS = {"Overall": "ภาพรวม", "Preform": "พรีฟอร์ม"}

# ---------- Helper functions (ORM + shared logic) ----------

def _is_staff_or_admin(user):
    if not user.is_authenticated:
        return False
    up = getattr(user, "userprofile", None)
    return (up and up.role in ["admin", "staff"]) or user.is_staff or user.is_superuser


def _annotate_lots(qs):
    """
    เพิ่มฟิลด์ produced_qty ให้แต่ละ Lot ด้วย ORM
    ใช้ sum ของ ScanRecord.qty เพื่อลด N+1 query
    """
    return qs.annotate(produced_qty=Coalesce(Sum("scans__qty"), 0))


def _build_lot_list(qs):
    """
    รับ queryset ของ Lot (ผ่านการ filter แล้ว) -> คืนค่า:
    - lots: list สำหรับใช้ใน template
    - summary: dict ค่า waiting / in_progress / finished / total_lots
    """
    qs = _annotate_lots(qs)

    lots = []
    waiting = 0
    in_progress = 0
    finished = 0

    for lot in qs.order_by("lot_no"):
        produced = lot.produced_qty or 0
        target = lot.target or lot.production_quantity or 0

        if target > 0:
            progress = min(100, int(produced * 100 / target))
        else:
            progress = 0

        if produced == 0:
            waiting += 1
        elif progress >= 100:
            finished += 1
        else:
            in_progress += 1

        boxes = 0
        if lot.pieces_per_box:
            boxes = int(produced / lot.pieces_per_box)

        lots.append(
            {
                "lot_no": lot.lot_no,
                "part_no": lot.part_no,
                "customer": lot.customer,
                "description": lot.description,
                "department": lot.department,
                "machine_no": lot.machine_no,
                "type": lot.type or "Order",
                "production_quantity": lot.production_quantity,
                "pieces_per_box": lot.pieces_per_box,
                "produced": produced,
                "target": target,
                "progress": progress,
                "boxes": boxes,
                "first_scan": lot.first_scan,
                "last_scan": lot.last_scan,
            }
        )

    summary = {
        "total_lots": qs.count(),
        "waiting": waiting,
        "in_progress": in_progress,
        "finished": finished,
    }
    return lots, summary


def _build_type_counts(qs):
    """
    นับจำนวน lot ตาม type ใช้แสดงกล่องด้านบนของ List View
    qs ควรเป็น queryset หลัง filter แผนก / search แต่ก่อน filter status
    """
    return {
        "all": qs.count(),
        "order": qs.filter(type__iexact="Order").count(),
        "sample": qs.filter(type__iexact="Sample").count(),
        "reserved": qs.filter(type__iexact="Reserved").count(),
        "extra": qs.filter(type__iexact="Extra").count(),
        "claim": qs.filter(type__iexact="Claim").count(),
    }

# ---------- Auth ----------
def login_page(request):
    if request.user.is_authenticated:
        return redirect("home_menu")

    if request.method == "POST":
        u = request.POST.get("username", "").strip()
        p = request.POST.get("password", "").strip()
        user = authenticate(request, username=u, password=p)
        if user:
            login(request, user)
            return redirect("home_menu")

        return render(
            request,
            "production/login.html",
            {"error": "ชื่อผู้ใช้หรือรหัสผ่านไม่ถูกต้อง"},
        )

    return render(request, "production/login.html")


def logout_view(request):
    logout(request)
    return redirect("login_page")


# ---------- Pages พื้นฐาน ----------


@login_required
def index(request):
    return render(request, "production/index.html")


@login_required
def department_select(request):
    return render(request, "production/department_select.html")


@login_required
def view_select(request):
    """
    เลือกมุมมองของแต่ละแผนก
    ถ้าเลือก Overall ให้ข้ามไป List View ทันที
    """
    dept = request.GET.get("department", "Overall")

    if dept.lower() == "overall":
        return redirect(f"{reverse('dashboard')}?department=Overall&view=list")

    return render(
        request,
        "production/view_select.html",
        {
            "department": dept,
            "department_label": LABELS.get(dept, dept),
        },
    )


# ---------- Dashboard หลัก (List / Machine / Order / Productivity) ----------


@login_required
def dashboard(request):
    dept = request.GET.get("department", "Overall")
    view_type = request.GET.get("view", "list")  # list / machine / order / productivity
    from_view = request.GET.get("from_view")     # ใช้ส่งต่อไปถึง lot_detail / dashboard_list

    # ---------- ถ้าเป็น Productivity ให้เด้งไปหน้าใหม่ทันที ----------
    if view_type == "productivity":
        from_date = request.GET.get("from", "")
        to_date = request.GET.get("to", "")

        url = reverse("productivity_form")  # ไปหน้าเลือกช่วงวันที่ก่อน
        params = [f"department={dept}"]
        if from_date:
            params.append(f"from={from_date}")
        if to_date:
            params.append(f"to={to_date}")

        return redirect(f"{url}?{'&'.join(params)}")
    # -------------------------------------------------------------------

    machine_no_filter = request.GET.get("machine_no", "").strip()
    lot_type = request.GET.get("lot_type", "all")  # ใช้กับปุ่ม filter ด้านบน
    layout = request.GET.get("layout", "cards")    # ใช้เปลี่ยน layout (order/machine)
    status = request.GET.get("status", "all")      # waiting / in_progress / finished

    # normalize layout
    if layout not in ["cards", "table"]:
        layout = "cards"

    department_label = LABELS.get(dept, dept)

    # ---------- ดึงข้อมูล Lot ----------
    qs = Lot.objects.all()

    # filter ตามแผนก
    if dept == "Preform":
        qs = qs.filter(department__icontains="พรีฟอร์ม")
    elif dept == "Overall":
        # ภาพรวมไม่ filter เพิ่ม
        pass
    else:
        qs = qs.filter(department__icontains=department_label)

    # filter ตามเครื่อง (ถ้ามาจาก machine_detail หรือ query)
    if machine_no_filter:
        qs = qs.filter(machine_no__iexact=machine_no_filter)

    # search
    q = request.GET.get("q", "").strip()
    if q:
        qs = qs.filter(
            Q(lot_no__icontains=q)
            | Q(part_no__icontains=q)
            | Q(customer__icontains=q)
        )

    # เก็บ qs เดิมไว้ใช้สรุป count ด้านบน (ไม่โดน filter lot_type / status)
    qs_for_counts = qs

    # ---------- filter ตาม type จากปุ่มด้านบน ----------
    active_type = lot_type
    if lot_type != "all":
        type_map = {
            "order": "Order",
            "sample": "Sample",
            "reserved": "Reserved",
            "extra": "Extra",
            "claim": "Claim",
        }
        t = type_map.get(lot_type.lower())
        if t:
            qs = qs.filter(type__iexact=t)

    # ---------- สร้าง list lots + summary ----------
    lots_all, summary = _build_lot_list(qs)

    # ---------- filter ตามสถานะ (ใช้เฉพาะ List View) ----------
    active_status = (
        status if status in ["all", "waiting", "in_progress", "finished"] else "all"
    )

    if view_type == "list" and active_status != "all":
        if active_status == "waiting":
            lots = [l for l in lots_all if l["produced"] == 0]
        elif active_status == "in_progress":
            lots = [l for l in lots_all if 0 < l["progress"] < 100]
        elif active_status == "finished":
            lots = [l for l in lots_all if l["progress"] >= 100]
        else:
            lots = lots_all
    else:
        lots = lots_all

    # ---------- นับจำนวน lot ตาม type (สำหรับกล่องด้านบน) ----------
    type_counts = _build_type_counts(qs_for_counts)

    # ---------- grouped_lots (ใช้กับ Order View แบบเดิม) ----------
    grouped_lots = {
        "Order": [], "Sample": [], "Reserved": [], "Extra": [], "Claim": [],
    }
    for lot in lots:
        t = (lot["type"] or "Order").title()
        if t not in grouped_lots:
            grouped_lots[t] = []
        grouped_lots[t].append(lot)

    # ---------- สรุปยอดรวมแบบ Order Dashboard ----------
    overall_qty_by_type = {
        "Order": (
            qs_for_counts.filter(type__iexact="Order")
            .aggregate(s=Sum("target"))["s"] or 0
        ),
        "Sample": (
            qs_for_counts.filter(type__iexact="Sample")
            .aggregate(s=Sum("target"))["s"] or 0
        ),
        "Reserved": (
            qs_for_counts.filter(type__iexact="Reserved")
            .aggregate(s=Sum("target"))["s"] or 0
        ),
        "Extra": (
            qs_for_counts.filter(type__iexact="Extra")
            .aggregate(s=Sum("target"))["s"] or 0
        ),
        "Claim": (
            qs_for_counts.filter(type__iexact="Claim")
            .aggregate(s=Sum("target"))["s"] or 0
        ),
    }
    overall_total_target = sum(overall_qty_by_type.values())

    # ------------------------------------------------------
    #  Machine summary สำหรับ Order View (การ์ดที่มีปุ่ม type ด้านล่าง)
    # ------------------------------------------------------
    machine_summaries = []
    machines = []

    if view_type == "order":
        # 1) ดึงข้อมูล Machine เพื่อเอา machine_type มาทำ label
        machine_qs = Machine.objects.all()
        if dept == "Preform":
            machine_qs = machine_qs.filter(department__icontains="พรีฟอร์ม")
        elif dept != "Overall":
            machine_qs = machine_qs.filter(department__icontains=department_label)

        machine_info = {
            m.machine_no: (m.machine_name or "เครื่องจักร")
            for m in machine_qs
        }

        machine_map = {}

        for lot in lots:
            machine_no = lot["machine_no"] or "ไม่ระบุเครื่อง"

            ms = machine_map.setdefault(
                machine_no,
                {
                    "machine_no": machine_no,
                    "machine_type_label": machine_info.get(machine_no, "เครื่องจักร"),
                    "total_target": 0,
                    "total_produced": 0,
                    "types": {
                        "Order":    {"target": 0, "count": 0},
                        "Sample":   {"target": 0, "count": 0},
                        "Reserved": {"target": 0, "count": 0},
                        "Extra":    {"target": 0, "count": 0},
                        "Claim":    {"target": 0, "count": 0},
                    },
                    "lots": [],
                },
            )

            ms["total_target"] += lot["target"]
            ms["total_produced"] += lot["produced"]

            t = (lot["type"] or "Order").title()
            if t not in ms["types"]:
                ms["types"][t] = {"target": 0, "count": 0}
            ms["types"][t]["target"] += lot["target"]
            ms["types"][t]["count"] += 1

            ms["lots"].append(lot)

        for ms in machine_map.values():
            if ms["total_target"] > 0:
                ms["progress"] = round(ms["total_produced"] * 100 / ms["total_target"])
            else:
                ms["progress"] = 0

        machine_summaries = list(machine_map.values())

    # ------------------------------------------------------
    #  Machine cards สำหรับ Machine View  (แสดงทุกเครื่องจาก Machine List)
    # ------------------------------------------------------
    if view_type == "machine":
        machine_map = {}

        # 1) รวมข้อมูลจาก lot ก่อน (เหมือนเดิม)
        for lot in lots:
            m_no = lot["machine_no"] or "-"
            info = machine_map.setdefault(
                m_no,
                {
                    "machine_no": m_no,
                    "lots": [],
                    "active_lot": None,
                    "status": "Ready",
                },
            )
            info["lots"].append(lot)

        # หา active lot + status จาก lot ที่มีอยู่
        for m_no, info in machine_map.items():
            running_lot = [x for x in info["lots"] if 0 < x["progress"] < 100]
            finished_lot = [x for x in info["lots"] if x["progress"] >= 100]

            if running_lot:
                active = sorted(
                    running_lot,
                    key=lambda x: x["last_scan"] or x["first_scan"] or datetime.min,
                )[-1]
                info["active_lot"] = active
                info["status"] = "Running"
            elif finished_lot:
                active = sorted(
                    finished_lot,
                    key=lambda x: x["last_scan"] or x["first_scan"] or datetime.min,
                )[-1]
                info["active_lot"] = active
                info["status"] = "Finished"
            elif info["lots"]:
                active = sorted(
                    info["lots"],
                    key=lambda x: x["last_scan"] or x["first_scan"] or datetime.min,
                )[-1]
                info["active_lot"] = active
                info["status"] = "Ready"
            else:
                info["active_lot"] = None
                info["status"] = "Ready"

        # 2) ดึงรายการเครื่องจากตาราง Machine แล้วเติมเครื่องที่ "ไม่มี lot" ให้ครบ
        master_qs = Machine.objects.all()
        if dept == "Preform":
            master_qs = master_qs.filter(department__icontains="พรีฟอร์ม")
        elif dept != "Overall":
            master_qs = master_qs.filter(department__icontains=department_label)

        for m in master_qs:
            m_no = m.machine_no or "-"
            if m_no not in machine_map:
                machine_map[m_no] = {
                    "machine_no": m_no,
                    "lots": [],
                    "active_lot": None,
                    "status": "Ready",   # ยังไม่มีงาน → Ready
                }

        # 3) แปลงเป็น list เรียงตามรหัสเครื่อง
        machines = sorted(
            machine_map.values(),
            key=lambda x: x["machine_no"] or "",
        )

    # ---------- เลือก template ----------
    template_map = {
        "list": "production/dashboard_list.html",
        "machine": "production/dashboard_machine.html",
        "order": "production/dashboard_order.html",
        "productivity": "production/dashboard_productivity.html",
    }
    template_name = template_map.get(view_type, "production/dashboard_list.html")

    context = {
        "department": dept,
        "department_label": department_label,
        "view_type": view_type,
        "type_counts": type_counts,
        "summary": summary,
        "lots": lots,
        "grouped_lots": grouped_lots,
        "search_query": q,
        "from_date": request.GET.get("from", ""),
        "to_date": request.GET.get("to", ""),
        "machine_no": machine_no_filter,
        "active_type": active_type,
        "active_status": active_status,
        "layout": layout,
        # สำหรับ Order View ใหม่
        "overall_qty_by_type": overall_qty_by_type,
        "overall_total_target": overall_total_target,
        "machine_summaries": machine_summaries,
        # สำหรับ Machine View
        "machines": machines,
        # ใช้ส่งต่อไป list → lot_detail
        "from_view": from_view,
    }
    return render(request, template_name, context)



# ---------- Machine detail (ใช้ template list เดิม) ----------

@login_required
def machine_detail(request, machine_no):
    """
    หน้าแสดง log การสแกนของ 'วันนี้' สำหรับเครื่องหนึ่งเครื่อง
    - filter ตาม machine_no
    - filter ตาม department (ถ้าไม่ใช่ Overall)
    - เรียงเวลาเก่า -> ใหม่
    """
    dept = request.GET.get("department", "Overall")
    department_label = LABELS.get(dept, dept)

    # วันนี้ (ตาม timezone ปัจจุบัน)
    today = timezone.localdate()
    start_dt = timezone.make_aware(
        datetime.combine(today, datetime.min.time())
    )
    end_dt = timezone.make_aware(
        datetime.combine(today, datetime.max.time())
    )

    # ดึง ScanRecord ของเครื่องนี้ในวันนี้
    scans_qs = ScanRecord.objects.filter(
        machine_no__iexact=machine_no,
        scanned_at__range=(start_dt, end_dt),
    ).select_related("lot").order_by("scanned_at")

    # filter ตามแผนก
    if dept == "Preform":
        scans_qs = scans_qs.filter(lot__department__icontains="พรีฟอร์ม")
    elif dept != "Overall":
        scans_qs = scans_qs.filter(lot__department__icontains=department_label)

    # สรุป
    total_today = scans_qs.aggregate(s=Sum("qty"))["s"] or 0
    first_scan = scans_qs.first()
    latest_scan = scans_qs.last()

    active_lot = latest_scan.lot if latest_scan else None
    target = 0
    produced = 0
    status = "Ready"

    if active_lot:
        target = active_lot.target or active_lot.production_quantity or 0
        produced = (
            ScanRecord.objects.filter(lot=active_lot).aggregate(s=Sum("qty"))["s"]
            or 0
        )
        if target and produced >= target:
            status = "Finished"
        elif produced > 0:
            status = "Running"

    # เตรียม row สำหรับตาราง
    rows = []
    for s in scans_qs:
        lot = s.lot
        local_dt = timezone.localtime(s.scanned_at)
        rows.append(
            {
                "time": local_dt.strftime("%H:%M:%S"),
                "lot_no": lot.lot_no if lot else "",
                "part_no": lot.part_no if lot else "",
                "customer": lot.customer if lot else "",
                "qty": s.qty or 0,
            }
        )

    context = {
        "department": dept,
        "department_label": department_label,
        "machine_no": machine_no,
        "today": today,
        "rows": rows,
        # summary card
        "status": status,
        "active_lot": active_lot,
        "target": target,
        "produced": produced,
        "total_today": total_today,
        "first_scan": first_scan.scanned_at if first_scan else None,
        "last_scan": latest_scan.scanned_at if latest_scan else None,
    }
    return render(request, "production/machine_detail.html", context)



# ---------- Lot detail + Chart ----------

@login_required
def lot_detail(request, lot_no):
    """
    หน้าแสดงรายละเอียด Lot + กราฟปริมาณการสแกน + ประวัติการสแกน
    - agg = hour/day/month ใช้กับกราฟ
    - scan_order = newest/oldest/qty_desc/qty_asc ใช้เรียงตารางประวัติ
    - scan_machine = all หรือรหัสเครื่อง
    - scan_from / scan_to = YYYY-MM-DD ใช้กรองช่วงวันที่
    """
    from datetime import datetime, timedelta
    from django.utils import timezone
    from django.db.models import Sum

    # ------------------ พารามิเตอร์พื้นฐาน ------------------
    dept_param = request.GET.get("department") or "Overall"

    back_view = request.GET.get("view") or "list"
    if back_view not in ["list", "order", "machine", "productivity"]:
        back_view = "list"

    lot_type = request.GET.get("lot_type", "")
    status = request.GET.get("status", "")
    search_q = request.GET.get("q", "")
    layout = request.GET.get("layout", "")

    machine_no = request.GET.get("machine_no", "").strip()
    from_view = request.GET.get("from_view", "").strip()

    # ------------------ คำนวณ back_url + back_label ------------------
    # กรณีที่มาจาก Machine View โดยตรง → กลับไปหน้า machine_detail
    if back_view == "machine" or from_view == "machine":
        if machine_no:
            base_url = reverse("machine_detail", args=[machine_no])
            params = [f"department={dept_param}"]
            if from_view:
                params.append(f"from_view={from_view}")
            back_url = f"{base_url}?{'&'.join(params)}"
        else:
            dash_url = reverse("dashboard")
            params = [f"department={dept_param}", "view=machine"]
            back_url = f"{dash_url}?{'&'.join(params)}"
        back_label = "กลับไป Machine View"

    else:
        # กรณีปกติ: กลับไปหน้า dashboard ตาม view เดิม (list / order / productivity)
        dash_url = reverse("dashboard")
        params = [f"department={dept_param}", f"view={back_view}"]

        if back_view in ["list", "order"]:
            if lot_type:
                params.append(f"lot_type={lot_type}")
            if status:
                params.append(f"status={status}")
            if search_q:
                params.append(f"q={search_q}")
            if machine_no:
                params.append(f"machine_no={machine_no}")
            if back_view == "order" and layout:
                params.append(f"layout={layout}")

        if from_view:
            params.append(f"from_view={from_view}")

        back_url = f"{dash_url}?{'&'.join(params)}"

        if back_view == "order":
            back_label = "กลับไป Order View"
        elif back_view == "productivity":
            back_label = "กลับไป Productivity"
        else:
            back_label = "กลับไป List View"

    # ------------------ พารามิเตอร์สำหรับกราฟ ------------------
    agg = request.GET.get("agg", "hour")
    if agg not in ["hour", "day", "month"]:
        agg = "hour"

    scan_order = request.GET.get("scan_order", "newest")
    scan_machine = request.GET.get("scan_machine", "all")
    scan_from = request.GET.get("scan_from", "").strip()
    scan_to = request.GET.get("scan_to", "").strip()

    lot = get_object_or_404(Lot, lot_no=lot_no)

    # ------------------ ข้อมูล Scan ทั้งหมดของ Lot ------------------
    scans_all = ScanRecord.objects.filter(lot=lot).order_by("scanned_at")

    # ถ้ายังไม่มีการสแกนเลย
    if not scans_all.exists():
        context = {
            "department":        dept_param,
            "department_label":  LABELS.get(dept_param, lot.department or dept_param),
            "lot":               lot,
            "produced":          0,
            "target":            lot.target or lot.production_quantity or 0,
            "progress":          0,
            "boxes":             0,
            "chart_labels":      [],
            "chart_daily":       [],
            "chart_cumulative":  [],
            "agg":               agg,
            "scan_logs":         [],
            "scan_order":        scan_order,
            "scan_machine":      scan_machine,
            "scan_from":         scan_from,
            "scan_to":           scan_to,
            "scan_machines":     [],
            "back_view":         back_view,
            "back_url":          back_url,
            "back_label":        back_label,
        }
        return render(request, "production/lot_detail.html", context)

    # ------------------ สรุปด้านบน ------------------
    produced = scans_all.aggregate(s=Sum("qty"))["s"] or 0   # ใช้ qty
    target = lot.target or lot.production_quantity or 0
    progress = round((produced / target) * 100, 1) if target > 0 else 0
    boxes = scans_all.count()

    # ------------------ queryset สำหรับกราฟ ------------------
    scans_for_chart = scans_all
    if scan_machine and scan_machine != "all":
        scans_for_chart = scans_for_chart.filter(machine_no__iexact=scan_machine)

    # filter ช่วงวันที่
    date_from = None
    date_to = None
    if scan_from:
        try:
            date_from = datetime.strptime(scan_from, "%Y-%m-%d").date()
        except ValueError:
            date_from = None
    if scan_to:
        try:
            date_to = datetime.strptime(scan_to, "%Y-%m-%d").date()
        except ValueError:
            date_to = None

    if date_from:
        scans_for_chart = scans_for_chart.filter(scanned_at__date__gte=date_from)
    if date_to:
        scans_for_chart = scans_for_chart.filter(scanned_at__date__lte=date_to)

    # ------------------ สร้างข้อมูลกราฟ ------------------
    chart_labels = []
    chart_daily = []
    chart_cumulative = []

    THAI_MONTH_ABBR = {
        1: "ม.ค.", 2: "ก.พ.", 3: "มี.ค.", 4: "เม.ย.",
        5: "พ.ค.", 6: "มิ.ย.", 7: "ก.ค.", 8: "ส.ค.",
        9: "ก.ย.", 10: "ต.ค.", 11: "พ.ย.", 12: "ธ.ค.",
    }

    if agg == "hour":
        # -------- รายชั่วโมง: โฟกัสแค่ 1 วัน (00:00–23:00) --------
        if date_from:
            focus_date = date_from
        else:
            focus_date = scans_all.latest("scanned_at").scanned_at.date()

        scans_day = scans_for_chart.filter(scanned_at__date=focus_date)

        qty_by_hour = {h: 0 for h in range(24)}
        for s in scans_day:
            local_dt = timezone.localtime(s.scanned_at)
            h = local_dt.hour
            qty_by_hour[h] += s.qty or 0      # ใช้ qty

        cumulative = 0
        for h in range(24):
            label = f"{h:02d}:00"
            qty = qty_by_hour[h]
            cumulative += qty

            chart_labels.append(label)
            chart_daily.append(qty)
            chart_cumulative.append(cumulative)

    elif agg == "day":
        # -------- รายวัน: label = "23 ม.ค." --------
        first_scan_date = scans_for_chart.first().scanned_at.date()
        last_scan_date = scans_for_chart.last().scanned_at.date()

        start_date = date_from or first_scan_date
        end_date = date_to or last_scan_date

        days = (end_date - start_date).days
        date_list = [start_date + timedelta(days=i) for i in range(days + 1)]

        qty_by_date = {d: 0 for d in date_list}
        for s in scans_for_chart:
            d = timezone.localtime(s.scanned_at).date()
            if d in qty_by_date:
                qty_by_date[d] += s.qty or 0   # ใช้ qty

        cumulative = 0
        for d in date_list:
            label = f"{d.day} {THAI_MONTH_ABBR[d.month]}"
            qty = qty_by_date[d]
            cumulative += qty

            chart_labels.append(label)
            chart_daily.append(qty)
            chart_cumulative.append(cumulative)

    else:
        # -------- รายเดือน: label = "ม.ค. 25" --------
        first_scan = scans_for_chart.first().scanned_at
        last_scan = scans_for_chart.last().scanned_at

        year_month_start = (first_scan.year, first_scan.month)
        year_month_end = (last_scan.year, last_scan.month)

        ym_list = []
        y, m = year_month_start
        while (y, m) <= year_month_end:
            ym_list.append((y, m))
            if m == 12:
                y += 1
                m = 1
            else:
                m += 1

        qty_by_month = {(y, m): 0 for (y, m) in ym_list}
        for s in scans_for_chart:
            dt = timezone.localtime(s.scanned_at)
            key = (dt.year, dt.month)
            if key in qty_by_month:
                qty_by_month[key] += s.qty or 0   # ใช้ qty

        cumulative = 0
        for (y, m) in ym_list:
            label = f"{THAI_MONTH_ABBR[m]} {str(y)[2:]}"
            qty = qty_by_month[(y, m)]
            cumulative += qty

            chart_labels.append(label)
            chart_daily.append(qty)
            chart_cumulative.append(cumulative)

    # ------------------ ตารางประวัติการสแกน ------------------
    scan_logs_qs = scans_all
    if scan_machine and scan_machine != "all":
        scan_logs_qs = scan_logs_qs.filter(machine_no__iexact=scan_machine)
    if date_from:
        scan_logs_qs = scan_logs_qs.filter(scanned_at__date__gte=date_from)
    if date_to:
        scan_logs_qs = scan_logs_qs.filter(scanned_at__date__lte=date_to)

    if scan_order == "oldest":
        scan_logs_qs = scan_logs_qs.order_by("scanned_at")
    elif scan_order == "qty_desc":
        scan_logs_qs = scan_logs_qs.order_by("-qty", "-scanned_at")
    elif scan_order == "qty_asc":
        scan_logs_qs = scan_logs_qs.order_by("qty", "-scanned_at")
    else:  # newest
        scan_logs_qs = scan_logs_qs.order_by("-scanned_at")

    scan_logs = list(scan_logs_qs)

    scan_machines = (
        scans_all.values_list("machine_no", flat=True)
        .order_by()
        .distinct()
    )

    # ------------------ render ------------------
    context = {
        "department":        dept_param,
        "department_label":  LABELS.get(dept_param, lot.department or dept_param),
        "lot":               lot,
        "produced":          produced,
        "target":            target,
        "progress":          progress,
        "boxes":             boxes,
        "chart_labels":      chart_labels,
        "chart_daily":       chart_daily,
        "chart_cumulative":  chart_cumulative,
        "agg":               agg,
        "scan_logs":         scan_logs,
        "scan_order":        scan_order,
        "scan_machine":      scan_machine,
        "scan_from":         scan_from,
        "scan_to":           scan_to,
        "scan_machines":     scan_machines,
        "back_view":         back_view,
        "back_url":          back_url,
        "back_label":        back_label,
    }
    return render(request, "production/lot_detail.html", context)


# ---------- หน้าประกอบอื่น ๆ ----------

@login_required
def productivity_form(request):
    """
    หน้าเลือกช่วงวันที่ -> พอกดสร้างรายงานจะ redirect ไปหน้า report จริง
    """
    dept = request.GET.get("department", "Preform")
    today = now().date()

    # ค่า default = วันนี้
    from_date = request.GET.get("from") or today.strftime("%Y-%m-%d")
    to_date = request.GET.get("to") or today.strftime("%Y-%m-%d")

    if request.method == "POST":
        from_date = request.POST.get("from") or from_date
        to_date = request.POST.get("to") or to_date

        # redirect ไปหน้า report ที่เราเคยทำ (productivity_view)
        url = reverse("productivity_view")
        qs = f"?department={dept}&from={from_date}&to={to_date}"
        return redirect(url + qs)

    context = {
        "department": dept,
        "department_label": LABELS.get(dept, dept),
        "from_date": from_date,
        "to_date": to_date,
    }
    return render(request, "production/productivity_form.html", context)


@login_required
def productivity_view(request):
    """
    Productivity Summary แบบรายเครื่อง + รายวัน
    เอาไอเดียมาจากหน้าเดิม (Netlify):
    - เลือกช่วงวันที่ (from / to)
    - รวมยอดผลิตต่อเครื่อง (sum ScanRecord.qty)
    - แสดงกราฟแท่ง + ตารางรายวัน + total
    """

    # -------- 1) อ่านค่าพื้นฐานจาก query string --------
    dept = request.GET.get("department", "Preform")
    department_label = LABELS.get(dept, dept)

    from_str = request.GET.get("from", "")
    to_str = request.GET.get("to", "")

    today = now().date()

    def parse_date(s):
        try:
            return datetime.strptime(s, "%Y-%m-%d").date()
        except (TypeError, ValueError):
            return None

    from_date = parse_date(from_str) or today
    to_date = parse_date(to_str) or today

    # ถ้าเลือก to น้อยกว่า from ให้บังคับให้เท่ากัน (กัน user ใส่ผิด)
    if to_date < from_date:
        to_date = from_date

    # สร้าง list ของวันที่ในช่วงนั้น (รวมปลายทาง)
    days = (to_date - from_date).days
    date_list = [from_date + timedelta(days=i) for i in range(days + 1)]

    # -------- 2) ดึง ScanRecord ในช่วงวันที่ + filter ตามแผนก --------
    scans = ScanRecord.objects.all()

    # filter ตามแผนก ใช้ department ของ Lot (เร็วและแม่นกว่า)
    if dept == "Preform":
        scans = scans.filter(lot__department__icontains="พรีฟอร์ม")
    elif dept != "Overall":
        scans = scans.filter(lot__department__icontains=department_label)

    # filter ช่วงวันที่ (ใช้ส่วน date ของ scanned_at)
    scans = scans.filter(
        scanned_at__date__gte=from_date,
        scanned_at__date__lte=to_date,
    )

    # -------- 3) รวมยอดต่อเครื่องต่อวันด้วย ORM --------
    grouped = (
        scans.annotate(day=TruncDate("scanned_at"))
        .values("machine_no", "day")
        .annotate(total_qty=Sum("qty"))
    )

    # เก็บเป็น dict โดย key = machine_no
    machine_data = {}  # {"M308": {"daily": {date1: qty, ...}, "total": sum}}
    for row in grouped:
        machine_no = row["machine_no"] or "-"
        day = row["day"]
        qty = row["total_qty"] or 0

        info = machine_data.setdefault(
            machine_no,
            {"machine_no": machine_no, "daily": {}, "total": 0},
        )
        info["daily"][day] = qty
        info["total"] += qty

    # -------- 4) ดึงชื่อเครื่องจากตาราง Machine (ถ้ามี) --------
    machine_rows = []
    if machine_data:
        machine_names = dict(
            Machine.objects.filter(machine_no__in=machine_data.keys()).values_list(
                "machine_no", "machine_name"
            )
        )

        # สร้าง list สำหรับ template
        for m_no in sorted(machine_data.keys()):
            info = machine_data[m_no]
            daily_list = [info["daily"].get(d, 0) for d in date_list]

            machine_rows.append(
                {
                    "machine_no": m_no,
                    "machine_name": machine_names.get(m_no, "") or "-",
                    "daily": daily_list,  # [qty_day1, qty_day2, ...]
                    "total": info["total"],  # รวมทุกวัน
                }
            )

    # -------- 5) สรุป total ต่อวัน + grand total --------
    total_per_day = []
    for idx, d in enumerate(date_list):
        total_per_day.append(sum(row["daily"][idx] for row in machine_rows))

    grand_total = sum(row["total"] for row in machine_rows)

    # -------- 6) เตรียมข้อมูลสำหรับกราฟ --------
    chart_labels = [row["machine_no"] for row in machine_rows]
    chart_data = [row["total"] for row in machine_rows]

    context = {
        "department": dept,
        "department_label": department_label,
        "from_date": from_date.strftime("%Y-%m-%d"),
        "to_date": to_date.strftime("%Y-%m-%d"),
        "date_list": date_list,
        "machine_rows": machine_rows,
        "total_per_day": total_per_day,
        "grand_total": grand_total,
        "chart_labels": chart_labels,
        "chart_data": chart_data,
    }
    return render(request, "production/productivity_view.html", context)


@login_required
def productivity_table(request):
    return render(request, "production/productivity_table.html")


@login_required
def scan(request):
    return render(request, "production/scan.html")


@login_required
def qr_export(request):
    return render(request, "production/qr_export.html")


@login_required
def user_profile(request):
    return render(request, "production/user_profile.html")


@login_required
@user_passes_test(_is_staff_or_admin)
def user_list_admin(request):
    return render(request, "production/user_list_admin.html")


@login_required
@user_passes_test(_is_staff_or_admin)
def data_collect(request):
    return render(request, "production/data_collect.html")


@login_required
@user_passes_test(_is_staff_or_admin)
def user_control(request):
    """
    หน้า User Control:
    - แสดงรายชื่อผู้ใช้ที่ออนไลน์ (จาก Session)
    - กดปุ่ม 'เตะออกจากระบบ' เพื่อลบ session ของ user นั้น
    """
    if request.method == "POST":
        action = request.POST.get("action")
        user_id = request.POST.get("user_id")

        if action == "kick" and user_id:
            try:
                _kick_user_sessions(int(user_id))
                messages.success(request, "เตะผู้ใช้ออกจากระบบเรียบร้อยแล้ว")
            except Exception as e:
                messages.error(request, f"ไม่สามารถเตะผู้ใช้ได้: {e}")

        # เตะเสร็จแล้วก็ redirect กลับหน้าเดิม (กันปัญหา refresh แล้วยิง POST ซ้ำ)
        from django.urls import reverse
        return redirect(reverse("user_control"))

    online_users = _get_online_users()

    context = {
        "online_users": online_users,
    }
    return render(request, "production/user_control.html", context)



# ---------- API (mock จากระบบเดิม) ----------


def _mock_if_empty():
    if not Lot.objects.exists():
        lot = Lot.objects.create(
            lot_no="LOT-AB-0001",
            part_no="PN-001",
            customer="A.Best",
            description="Demo lot",
            production_quantity=1000,
            pieces_per_box=50,
            target=1000,
            department="Assembly",
            machine_no="MC-01",
            type="Normal",
            first_scan=now(),
            last_scan=now(),
        )
        ScanRecord.objects.create(lot=lot, machine_no="MC-01", qty=250)
        ScanRecord.objects.create(lot=lot, machine_no="MC-01", qty=300)


@csrf_exempt
def api(request):
    action = request.POST.get("action") or request.GET.get("action")
    if not action:
        return JsonResponse(
            {"status": "error", "message": "Missing action"}, status=400
        )

    # mock login
    if action == "login":
        user = request.POST.get("user") or request.GET.get("user")
        password = request.POST.get("password") or request.GET.get("password")
        ok = (user in ["admin", "staff", "visitor"]) and (password == "1234")
        if ok:
            return JsonResponse(
                {
                    "status": "success",
                    "user": {
                        "name": user.title(),
                        "username": user,
                        "role": user,
                    },
                    "token": "demo-token",
                }
            )
        return JsonResponse(
            {"status": "error", "message": "Invalid credentials"}, status=401
        )

    if action == "getData":
        _mock_if_empty()
        rows = []
        for lot in Lot.objects.all().order_by("lot_no"):
            produced = lot.scans.aggregate(s=Sum("qty"))["s"] or 0
            progress = (
                0
                if not lot.target
                else min(100, int(produced * 100 / lot.target))
            )
            rows.append(
                {
                    "lotNo": lot.lot_no,
                    "partNo": lot.part_no,
                    "customer": lot.customer,
                    "description": lot.description,
                    "productionQuantity": lot.production_quantity,
                    "piecesPerBox": lot.pieces_per_box,
                    "target": lot.target,
                    "department": lot.department,
                    "machineNo": lot.machine_no,
                    "type": lot.type or "Order",
                    "firstScan": lot.first_scan.isoformat()
                    if lot.first_scan
                    else None,
                    "lastScan": lot.last_scan.isoformat()
                    if lot.last_scan
                    else None,
                    "scannedCount": produced,
                    "progress": progress,
                }
            )
        return JsonResponse(
            {
                "status": "success",
                "data": {
                    "dashboardData": rows,
                    "machineData": [],
                    "scanLog": [],
                    "orderViewSummary": {},
                },
            }
        )

    if action == "scan":
        lot_no = request.POST.get("lot_no") or request.GET.get("lot_no")
        qty = int(request.POST.get("qty") or request.GET.get("qty") or 0)
        machine_no = (
            request.POST.get("machine_no")
            or request.GET.get("machine_no")
            or "MC-01"
        )
        try:
            lot = Lot.objects.get(lot_no=lot_no)
        except Lot.DoesNotExist:
            return JsonResponse(
                {
                    "status": "error",
                    "message": f"Lot {lot_no} not found",
                },
                status=404,
            )
        ScanRecord.objects.create(lot=lot, machine_no=machine_no, qty=qty)
        lot.last_scan = now()
        lot.first_scan = lot.first_scan or lot.last_scan
        lot.save(update_fields=["first_scan", "last_scan"])
        return JsonResponse({"status": "success"})

    if action in ["getActiveUsers", "kickUser", "getQrExportData"]:
        return JsonResponse({"status": "success", "data": []})

    return JsonResponse(
        {"status": "error", "message": "Unknown action"}, status=400
    )


# ---------- Dashboard shortcuts (Overall / Preform) ----------


@login_required
def dashboard_overview(request):
    """
    Shortcut: Overall – List View
    /dashboard/overview/ -> /dashboard/?department=Overall&view=list
    """
    url = reverse("dashboard")
    return redirect(f"{url}?department=Overall&view=list")


@login_required
def dashboard_preform(request):
    """
    Shortcut: Preform – List View
    /dashboard/preform/ -> /dashboard/?department=Preform&view=list
    """
    url = reverse("dashboard")
    return redirect(f"{url}?department=Preform&view=list")


@login_required
def dashboard_overall_order(request):
    """
    Shortcut: Overall – Order View
    /dashboard/overall/order/ -> /dashboard/?department=Overall&view=order
    """
    url = reverse("dashboard")
    return redirect(f"{url}?department=Overall&view=order")


@login_required
def dashboard_preform_order(request):
    """
    Shortcut: Preform – Order View
    /dashboard/preform/order/ -> /dashboard/?department=Preform&view=order
    """
    url = reverse("dashboard")
    return redirect(f"{url}?department=Preform&view=order")


@login_required
def export_productivity_excel(request):
    # 1) รับ filter (เหมือนหน้า Dashboard)
    dept = request.GET.get("department", "Overall")
    machine_no_filter = request.GET.get("machine_no", "").strip()
    date_from = request.GET.get("from", "")
    date_to = request.GET.get("to", "")

    # 2) Query ข้อมูลพื้นฐาน
    qs = Lot.objects.all()

    # Filter แผนก
    if dept == "Preform":
        qs = qs.filter(department__icontains="พรีฟอร์ม")
    elif dept != "Overall":
        qs = qs.filter(department__icontains=LABELS.get(dept, dept))

    # Filter เครื่อง
    if machine_no_filter:
        qs = qs.filter(machine_no__iexact=machine_no_filter)

    # Filter วันที่
    if date_from:
        qs = qs.filter(last_scan__date__gte=date_from)
    if date_to:
        qs = qs.filter(last_scan__date__lte=date_to)

    # Annotate ผลรวมการผลิต (ORM ลด Query)
    qs = _annotate_lots(qs).order_by("lot_no")

    # 3) สร้าง Excel Workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Productivity Report"

    headers = [
        "Lot No",
        "Part No",
        "Customer",
        "Department",
        "Machine",
        "Type",
        "Target",
        "Produced",
        "Progress (%)",
        "Boxes",
        "Status",
        "Last Scan",
    ]
    ws.append(headers)

    # 4) เติมข้อมูลทีละแถว
    for lot in qs:
        produced = lot.produced_qty or 0
        target = lot.target or lot.production_quantity or 0
        progress = (produced / target * 100) if target > 0 else 0

        # กล่อง
        boxes = int(produced / lot.pieces_per_box) if lot.pieces_per_box else 0

        # สถานะ
        if produced == 0:
            status = "Waiting"
        elif progress >= 100:
            status = "Finished"
        else:
            status = "Running"

        # เวลา Scan
        last_scan_str = (
            lot.last_scan.strftime("%Y-%m-%d %H:%M") if lot.last_scan else ""
        )

        ws.append(
            [
                lot.lot_no,
                lot.part_no,
                lot.customer,
                lot.department,
                lot.machine_no,
                lot.type or "Order",
                target,
                produced,
                round(progress, 2),
                boxes,
                status,
                last_scan_str,
            ]
        )

    # 5) ปรับความกว้างคอลัมน์ให้พอดี
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 16

    # 6) ส่งกลับเป็นไฟล์ดาวน์โหลด
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    file_name = f"Productivity_Report_{now().strftime('%Y%m%d')}.xlsx"
    response["Content-Disposition"] = f'attachment; filename="{file_name}"'

    wb.save(response)
    return response


@login_required
def import_excel(request):
    if request.method == "POST" and request.FILES.get("excel_file"):
        excel_file = request.FILES["excel_file"]
        filename = excel_file.name.lower()

        # รองรับทั้ง Excel และ CSV
        if not filename.endswith((".xls", ".xlsx", ".csv")):
            messages.error(
                request, "กรุณาอัปโหลดไฟล์ Excel (.xlsx) หรือ CSV (.csv)"
            )
            return redirect("import_excel")

        try:
            # อ่านไฟล์เป็น DataFrame
            if filename.endswith(".csv"):
                df = pd.read_csv(excel_file)
            else:
                df = pd.read_excel(excel_file)

            # เคลียร์ช่องว่างในชื่อคอลัมน์
            df.columns = df.columns.str.strip()

            count = 0

            # ---------- กรณี 1: นำเข้า Machine List ----------
            if "Machine Name" in df.columns and "Machine No." in df.columns:
                for _, row in df.iterrows():
                    if pd.isna(row.get("Machine No.")):
                        continue

                    Machine.objects.update_or_create(
                        machine_no=str(row["Machine No."]).strip(),
                        defaults={
                            "machine_name": str(
                                row.get("Machine Name", "")
                            ).strip(),
                            "department": str(
                                row.get("Department", "")
                            ).strip(),
                        },
                    )
                    count += 1
                messages.success(
                    request, f"นำเข้าข้อมูลเครื่องจักรสำเร็จ {count} รายการ"
                )

            # ---------- กรณี 2: นำเข้า Lot / Databased ----------
            elif "Lot No." in df.columns:
                for _, row in df.iterrows():
                    if pd.isna(row.get("Lot No.")):
                        continue

                    prod_qty = int(
                        pd.to_numeric(
                            row.get("Production Quantity"), errors="coerce"
                        )
                        or 0
                    )
                    pieces_per_box = int(
                        pd.to_numeric(
                            row.get("จำนวนบรรจุต่อกล่อง"), errors="coerce"
                        )
                        or 0
                    )

                    Lot.objects.update_or_create(
                        lot_no=str(row["Lot No."]).strip(),
                        defaults={
                            "part_no": str(
                                row.get("A.Best Part No.", "")
                            ).strip(),
                            "customer": str(row.get("Customer", "")).strip(),
                            "description": str(
                                row.get("Description", "")
                            ).strip(),
                            "customer_part_no": str(
                                row.get("Customer Part No.", "")
                            ).strip(),
                            "po_no": str(row.get("PO No.", "")).strip(),
                            "remark": str(row.get("Remark", "")).strip(),
                            "production_quantity": prod_qty,
                            "target": prod_qty,
                            "pieces_per_box": pieces_per_box,
                            "department": str(
                                row.get("Department", "Overall")
                            ).strip(),
                            "machine_no": str(
                                row.get("Machine No.", "")
                            ).strip(),
                            "type": str(row.get("Type", "Order")).strip(),
                        },
                    )
                    count += 1

                messages.success(
                    request, f"นำเข้าแผนการผลิตสำเร็จ {count} รายการ"
                )

            else:
                messages.error(
                    request,
                    "ไม่พบรูปแบบข้อมูลที่รองรับ (ต้องเป็นไฟล์ Databased หรือ Machine List)",
                )

        except Exception as e:
            messages.error(request, f"เกิดข้อผิดพลาด: {e}")

        return redirect("dashboard")

    return render(request, "production/import_excel.html")


class Command(BaseCommand):
    help = "Import data from 'A.Best - Production Tracker.xlsx' into Django models"

    def add_arguments(self, parser):
        parser.add_argument(
            "path",
            nargs="?",
            default="A.Best - Production Tracker.xlsx",
            help="Path to the Excel file",
        )

    def handle(self, *args, **options):
        path = options["path"]
        self.stdout.write(self.style.NOTICE(f"Loading workbook: {path}"))
        wb = openpyxl.load_workbook(path, data_only=True)

        with transaction.atomic():
            self._import_machines(wb)
            lot_map = self._import_lots(wb)
            self._import_collect(wb, lot_map)

        self.stdout.write(self.style.SUCCESS("Import completed."))

    # ------------------------ Machines ------------------------

    def _import_machines(self, wb):
        if "Machine List" not in wb.sheetnames:
            self.stdout.write("Sheet 'Machine List' not found, skip machines.")
            return

        ws = wb["Machine List"]
        headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        idx = {name: i for i, name in enumerate(headers) if name}

        required = ["Machine No.", "Machine Name", "Department"]
        for r in required:
            if r not in idx:
                self.stdout.write(
                    self.style.WARNING(f"Column '{r}' not found in Machine List")
                )

        count = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            machine_no = row[idx.get("Machine No.")] if "Machine No." in idx else None
            if not machine_no:
                continue
            machine_name = (
                row[idx.get("Machine Name")] if "Machine Name" in idx else ""
            )
            department = row[idx.get("Department")] if "Department" in idx else ""

            Machine.objects.update_or_create(
                machine_no=str(machine_no).strip(),
                defaults={
                    "machine_name": str(machine_name or "").strip(),
                    "department": str(department or "").strip(),
                },
            )
            count += 1
        self.stdout.write(self.style.SUCCESS(f"Imported/updated {count} machines."))

    # ------------------------ Lots (Databased) ------------------------

    def _import_lots(self, wb):
        if "Databased" not in wb.sheetnames:
            self.stdout.write("Sheet 'Databased' not found, skip lots.")
            return {}

        ws = wb["Databased"]
        headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        idx = {name: i for i, name in enumerate(headers) if name}

        lot_map = {}
        count = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            lot_no = row[idx.get("Lot No.")] if "Lot No." in idx else None
            if not lot_no:
                continue

            abest_part_no = (
                row[idx.get("A.Best Part No.")] if "A.Best Part No." in idx else ""
            )
            customer = row[idx.get("Customer")] if "Customer" in idx else ""
            description = row[idx.get("Description")] if "Description" in idx else ""
            customer_part_no = (
                row[idx.get("Customer Part No.")] if "Customer Part No." in idx else ""
            )
            po_no = row[idx.get("PO No.")] if "PO No." in idx else ""
            remark = row[idx.get("Remark")] if "Remark" in idx else ""
            department = row[idx.get("Department")] if "Department" in idx else ""
            machine_no = row[idx.get("Machine No.")] if "Machine No." in idx else ""
            lot_type = row[idx.get("Type")] if "Type" in idx else "Order"

            prod_qty = (
                row[idx.get("Production Quantity")]
                if "Production Quantity" in idx
                else 0
            )
            pieces_per_box = (
                row[idx.get("จำนวนบรรจุต่อกล่อง")]
                if "จำนวนบรรจุต่อกล่อง" in idx
                else 0
            )
            target = row[idx.get("Target")] if "Target" in idx else None

            try:
                prod_qty = int(prod_qty or 0)
            except Exception:
                prod_qty = 0
            try:
                pieces_per_box = int(pieces_per_box or 0)
            except Exception:
                pieces_per_box = 0
            try:
                target = int(target) if target is not None else prod_qty
            except Exception:
                target = prod_qty

            lot, created = Lot.objects.update_or_create(
                lot_no=str(lot_no).strip(),
                defaults={
                    "part_no": str(abest_part_no or "").strip(),
                    "customer": str(customer or "").strip(),
                    "description": str(description or "").strip(),
                    "customer_part_no": str(customer_part_no or "").strip(),
                    "po_no": str(po_no or "").strip(),
                    "remark": str(remark or "").strip(),
                    "production_quantity": prod_qty,
                    "pieces_per_box": pieces_per_box,
                    "target": target,
                    "department": str(department or "").strip(),
                    "machine_no": str(machine_no or "").strip(),
                    "type": str(lot_type or "Order").strip(),
                },
            )
            lot_map[lot.lot_no] = lot
            count += 1

        self.stdout.write(self.style.SUCCESS(f"Imported/updated {count} lots."))
        return lot_map

    # ------------------------ Scan Records (Collect) ------------------------

    def _import_collect(self, wb, lot_map):
        if "Collect" not in wb.sheetnames:
            self.stdout.write("Sheet 'Collect' not found, skip scan records.")
            return

        ws = wb["Collect"]
        headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        idx = {name: i for i, name in enumerate(headers) if name}

        count = 0
        per_lot_first = {}
        per_lot_last = {}

        for row in ws.iter_rows(min_row=2, values_only=True):
            lot_no = row[idx.get("Lot No.")] if "Lot No." in idx else None
            if not lot_no:
                continue

            lot_no = str(lot_no).strip()
            lot = lot_map.get(lot_no) or Lot.objects.filter(lot_no=lot_no).first()
            if not lot:
                # ถ้าไม่มี lot ใน Databased ข้ามไป
                continue

            date_val = row[idx.get("Date")] if "Date" in idx else None
            time_val = row[idx.get("Time")] if "Time" in idx else None
            dept = row[idx.get("Department")] if "Department" in idx else ""
            machine_no = row[idx.get("Machine No.")] if "Machine No." in idx else ""

            if date_val is None:
                continue

            # date_val เป็น datetime หรือ date, time_val เป็น time
            if isinstance(date_val, datetime):
                d = date_val.date()
            else:
                d = date_val
            if time_val is None:
                scan_dt = datetime.combine(d, datetime.min.time())
            else:
                scan_dt = datetime.combine(d, time_val)

            qty = lot.pieces_per_box or 0  # 1 scan = 1 กล่อง

            scan, created = ScanRecord.objects.get_or_create(
                lot=lot,
                machine_no=str(machine_no or "").strip(),
                scanned_at=scan_dt,
                defaults={"qty": qty},
            )
            if not created:
                # ถ้าเคยมีอยู่แล้ว ไม่ต้องทำซ้ำ
                continue

            # เก็บ first/last scan per lot
            per_lot_first[lot.pk] = min(
                per_lot_first.get(lot.pk, scan_dt), scan_dt
            )
            per_lot_last[lot.pk] = max(
                per_lot_last.get(lot.pk, scan_dt), scan_dt
            )

            count += 1

        # อัปเดต first_scan / last_scan
        for lot_id, lot in Lot.objects.in_bulk(per_lot_first.keys()).items():
            lot.first_scan = per_lot_first.get(lot_id)
            lot.last_scan = per_lot_last.get(lot_id)
            lot.save(update_fields=["first_scan", "last_scan"])

        self.stdout.write(
            self.style.SUCCESS(f"Imported {count} scan records from Collect.")
        )


@login_required
def lot_chart_data(request, lot_no):
    """
    คืนค่า labels / daily / cumulative เป็น JSON สำหรับกราฟใน lot_detail

    พฤติกรรม:
    - agg=hour  -> เฉพาะ 1 วัน (00–23) ใช้ param ?date=YYYY-MM-DD ถ้ามี
    - agg=day   -> รายวัน, label "23 ก.ย." รองรับ ?from / ?to (YYYY-MM-DD)
    - agg=month -> รายเดือน, label "ก.ย. 25" (ปี 2 หลัก) รองรับ ?from / ?to
    """
    from django.utils import timezone  # ใช้แปลงเวลาเป็น localtime

    agg = request.GET.get("agg", "hour")
    if agg not in ["hour", "day", "month"]:
        agg = "hour"

    lot = get_object_or_404(Lot, lot_no=lot_no)

    # ข้อมูลสแกนทั้งหมดของ Lot นี้
    scans_all = ScanRecord.objects.filter(lot=lot).order_by("scanned_at")

    if not scans_all.exists():
        return JsonResponse({"labels": [], "daily": [], "cumulative": []})

    # helper แปลง string -> date
    def parse_date(s):
        try:
            return datetime.strptime(s, "%Y-%m-%d").date()
        except (TypeError, ValueError):
            return None

    # --- อ่านช่วงวันที่จาก query (ใช้กับ day / month, และ hour แบบ optional) ---
    date_from = parse_date(request.GET.get("from"))
    date_to = parse_date(request.GET.get("to"))
    focus_date = parse_date(request.GET.get("date"))

    labels = []
    daily = []
    cumulative = []
    running = 0

    # ---------- โหมดรายชั่วโมง (00–23 ของวันเดียว) ----------
    if agg == "hour":
        # วันเป้าหมาย:
        # 1) ถ้ามี ?date=... ใช้วันนั้น
        # 2) ถ้าไม่มีก็ใช้วันที่ของ record ล่าสุด
        if focus_date:
            target_date = focus_date
        else:
            target_date = timezone.localtime(
                scans_all.latest("scanned_at").scanned_at
            ).date()

        scans_day = scans_all.filter(scanned_at__date=target_date)

        # เตรียม dict 24 ชั่วโมงให้ครบก่อน (แม้จะไม่มี data ก็เป็น 0)
        qty_by_hour = {h: 0 for h in range(24)}
        for s in scans_day:
            local_dt = timezone.localtime(s.scanned_at)
            h = local_dt.hour
            qty_by_hour[h] += s.qty or 0  # ใช้ qty

        for h in range(24):
            q = qty_by_hour[h]
            running += q
            labels.append(f"{h:02d}:00")
            daily.append(q)
            cumulative.append(running)

    # ---------- โหมดรายวัน ----------
    elif agg == "day":
        # ช่วงวัน: ใช้ from/to ถ้ามี ไม่งั้นใช้ first/last scan
        first_date = timezone.localtime(scans_all.first().scanned_at).date()
        last_date = timezone.localtime(scans_all.last().scanned_at).date()

        start_date = date_from or first_date
        end_date = date_to or last_date
        if end_date < start_date:
            end_date = start_date

        days = (end_date - start_date).days
        date_list = [start_date + timedelta(days=i) for i in range(days + 1)]

        qty_by_date = {d: 0 for d in date_list}
        for s in scans_all:
            d = timezone.localtime(s.scanned_at).date()
            if d in qty_by_date:
                qty_by_date[d] += s.qty or 0  # ใช้ qty

        for d in date_list:
            q = qty_by_date[d]
            running += q
            labels.append(f"{d.day} {MONTH_TH[d.month]}")
            daily.append(q)
            cumulative.append(running)

    # ---------- โหมดรายเดือน ----------
    else:  # agg == "month"
        first_dt = timezone.localtime(scans_all.first().scanned_at)
        last_dt = timezone.localtime(scans_all.last().scanned_at)

        # ให้ from/to (ถ้ามี) ครอบช่วงเดือน
        if date_from:
            first_dt = first_dt.replace(year=date_from.year, month=date_from.month, day=1)
        if date_to:
            last_dt = last_dt.replace(year=date_to.year, month=date_to.month, day=1)

        start_ym = (first_dt.year, first_dt.month)
        end_ym = (last_dt.year, last_dt.month)

        ym_list = []
        y, m = start_ym
        while (y, m) <= end_ym:
            ym_list.append((y, m))
            if m == 12:
                y += 1
                m = 1
            else:
                m += 1

        qty_by_month = {(y, m): 0 for (y, m) in ym_list}
        for s in scans_all:
            dt = timezone.localtime(s.scanned_at)
            key = (dt.year, dt.month)
            if key in qty_by_month:
                qty_by_month[key] += s.qty or 0  # ใช้ qty

        for (y, m) in ym_list:
            q = qty_by_month[(y, m)]
            running += q
            # ปี 2 หลัก เช่น 2025 -> "25"
            labels.append(f"{MONTH_TH[m]} {str(y)[2:]}")
            daily.append(q)
            cumulative.append(running)

    return JsonResponse(
        {"labels": labels, "daily": daily, "cumulative": cumulative}
    )
    
# ---------- Helper สำหรับ User Control (ออนไลน์ / เตะออก) ----------

def _get_online_users():
    """
    คืนค่า list ของผู้ใช้ที่กำลังออนไลน์ จากตาราง Session
    แต่ละ item เป็น dict: {
        "user": User object,
        "username": str,
        "full_name": str,
        "role": str (จาก UserProfile ถ้ามี),
        "department": str (จาก UserProfile ถ้ามี),
        "last_login": datetime | None,
    }
    """
    sessions = Session.objects.filter(expire_date__gt=timezone.now())
    user_ids = set()

    for s in sessions:
        data = s.get_decoded()
        uid = data.get("_auth_user_id")
        if uid:
            user_ids.add(int(uid))

    if not user_ids:
        return []

    users = (
        User.objects.filter(id__in=user_ids)
        .select_related("userprofile")
        .order_by("username")
    )

    online = []
    for u in users:
        profile = getattr(u, "userprofile", None)
        full_name = (u.get_full_name() or u.username).strip()
        role = profile.role if profile else "-"
        dept = profile.department.name if (profile and profile.department) else "-"

        online.append(
            {
                "user": u,
                "username": u.username,
                "full_name": full_name or u.username,
                "role": role,
                "department": dept,
                "last_login": u.last_login,
            }
        )
    return online


def _kick_user_sessions(user_id: int):
    """
    ลบทุก session ของ user คนนี้ -> ทำให้ user นั้นหลุดออกจากระบบทันที
    """
    sessions = Session.objects.filter(expire_date__gt=timezone.now())
    for s in sessions:
        data = s.get_decoded()
        if data.get("_auth_user_id") == str(user_id):
            s.delete()

@login_required
def machine_mini_chart(request, machine_no):
    """
    คืนข้อมูลกราฟ mini chart ของแต่ละเครื่อง (รายชั่วโมงของวันนี้)
    response: { "labels": [...], "daily": [...] }
    """
    dept = request.GET.get("department", "Overall")

    now = timezone.localtime(timezone.now())
    start_of_day = now.replace(hour=0, minute=0, second=0, microsecond=0)
    end_of_day = start_of_day.replace(hour=23, minute=59, second=59)

    scans = ScanRecord.objects.filter(
        machine_no__iexact=machine_no,
        scanned_at__range=(start_of_day, end_of_day),
    )

    if dept != "Overall":
        scans = scans.filter(lot__department__icontains=dept)

    grouped = (
        scans.annotate(h=TruncHour("scanned_at"))
        .values("h")
        .annotate(total_qty=Sum("qty"))
        .order_by("h")
    )

    qty_by_hour = {h: 0 for h in range(24)}
    for row in grouped:
        h = timezone.localtime(row["h"]).hour
        qty_by_hour[h] = row["total_qty"] or 0

    labels = [f"{h:02d}:00" for h in range(24)]
    daily = [qty_by_hour[h] for h in range(24)]

    return JsonResponse({"labels": labels, "daily": daily})


@login_required
def machine_chart_data(request, machine_no):
    """
    คืนค่า JSON สรุปข้อมูลเครื่อง + กราฟยอดสแกนรายชั่วโมงของวันนี้
    ใช้กับการ์ดใน Machine View
    """
    today = timezone.localdate()

    scans_qs = ScanRecord.objects.filter(
        machine_no__iexact=machine_no,
        scanned_at__date=today,
    ).order_by("scanned_at")

    qty_by_hour = {h: 0 for h in range(24)}
    latest_scan = None
    for s in scans_qs:
        local_dt = timezone.localtime(s.scanned_at)
        h = local_dt.hour
        qty_by_hour[h] += s.qty or 0
        latest_scan = s

    labels = [f"{h:02d}:00" for h in range(24)]
    daily = [qty_by_hour[h] for h in range(24)]

    lot = None
    if latest_scan:
        lot = latest_scan.lot
    else:
        latest_scan_all = (
            ScanRecord.objects.filter(machine_no__iexact=machine_no)
            .order_by("scanned_at")
            .last()
        )
        if latest_scan_all:
            latest_scan = latest_scan_all
            lot = latest_scan_all.lot

    target = 0
    produced = 0
    lot_no = ""
    part_no = ""
    customer = ""
    last_scan_display = ""

    if lot:
        lot_no = lot.lot_no or ""
        part_no = lot.part_no or ""
        customer = lot.customer or ""
        target = lot.target or lot.production_quantity or 0
        produced = (
            ScanRecord.objects.filter(lot=lot).aggregate(s=Sum("qty"))["s"] or 0
        )

    if latest_scan:
        last_scan_display = timezone.localtime(
            latest_scan.scanned_at
        ).strftime("%d/%m %H:%M")

    if target and produced >= target:
        status = "Finished"
    elif produced > 0:
        status = "Running"
    else:
        status = "Ready"

    data = {
        "machine_no": machine_no,
        "status": status,
        "lot_no": lot_no,
        "part_no": part_no,
        "customer": customer,
        "target": int(target),
        "produced": int(produced),
        "last_scan_display": last_scan_display,
        "labels": labels,
        "daily": daily,
    }
    return JsonResponse(data)

@login_required
def machine_chart_data(request, machine_no):
    """
    คืนค่า JSON สรุปข้อมูลเครื่อง + กราฟยอดสแกนรายชั่วโมงของ
    'วันล่าสุดที่มีการสแกนเครื่องนี้'
    ใช้กับการ์ดใน Machine View
    """
    # หา log ล่าสุดของเครื่องนี้ก่อน
    latest_scan_all = (
        ScanRecord.objects
        .filter(machine_no__iexact=machine_no)
        .order_by("scanned_at")
        .last()
    )

    if not latest_scan_all:
        # ยังไม่เคยสแกนเลย → คืนค่าเปล่า ๆ
        return JsonResponse({
            "machine_no": machine_no,
            "status": "Ready",
            "lot_no": "",
            "part_no": "",
            "customer": "",
            "target": 0,
            "produced": 0,
            "last_scan_display": "",
            "labels": [f"{h:02d}:00" for h in range(24)],
            "daily": [0 for _ in range(24)],
        })

    # ใช้ "วันที่ของ log ล่าสุด" เป็นวันเป้าหมายของกราฟ
    focus_date = timezone.localtime(latest_scan_all.scanned_at).date()

    scans_qs = (
        ScanRecord.objects
        .filter(
            machine_no__iexact=machine_no,
            scanned_at__date=focus_date,
        )
        .order_by("scanned_at")
    )

    # รวมจำนวนสแกนต่อชั่วโมง
    qty_by_hour = {h: 0 for h in range(24)}
    latest_scan = None
    for s in scans_qs:
        local_dt = timezone.localtime(s.scanned_at)
        h = local_dt.hour
        qty_by_hour[h] += s.qty or 0
        latest_scan = s

    labels = [f"{h:02d}:00" for h in range(24)]
    daily = [qty_by_hour[h] for h in range(24)]

    # lot ล่าสุด (จาก log ล่าสุดของทั้งเครื่อง)
    lot = latest_scan_all.lot if latest_scan_all else None

    target = 0
    produced = 0
    lot_no = ""
    part_no = ""
    customer = ""
    last_scan_display = ""

    if lot:
        lot_no = lot.lot_no or ""
        part_no = lot.part_no or ""
        customer = lot.customer or ""
        target = lot.target or lot.production_quantity or 0
        produced = (
            ScanRecord.objects.filter(lot=lot).aggregate(s=Sum("qty"))["s"] or 0
        )

    if latest_scan_all:
        last_scan_display = timezone.localtime(
            latest_scan_all.scanned_at
        ).strftime("%d/%m %H:%M")

    # สถานะเครื่องแบบง่าย ๆ
    if target and produced >= target:
        status = "Finished"
    elif produced > 0:
        status = "Running"
    else:
        status = "Ready"

    data = {
        "machine_no": machine_no,
        "status": status,
        "lot_no": lot_no,
        "part_no": part_no,
        "customer": customer,
        "target": int(target),
        "produced": int(produced),
        "last_scan_display": last_scan_display,
        "labels": labels,
        "daily": daily,
    }
    return JsonResponse(data)

@login_required
def machine_scan_logs_today(request, machine_no):
    today = timezone.localdate()

    scans = (
        ScanRecord.objects
        .filter(
            machine_no__iexact=machine_no,
            scanned_at__date=today
        )
        .select_related("lot")
        .order_by("-scanned_at")
    )

    data = []
    total_qty = 0

    for s in scans:
        lot = s.lot
        data.append({
            "time": timezone.localtime(s.scanned_at).strftime("%H:%M"),
            "lot_no": lot.lot_no if lot else "-",
            "part_no": lot.part_no if lot else "-",
            "customer": lot.customer if lot else "-",
            "qty": s.qty or 0,
        })
        total_qty += s.qty or 0

    return JsonResponse({"logs": data, "total": total_qty})

def _is_admin(user):
    if not user.is_authenticated:
        return False
    profile = getattr(user, "userprofile", None)
    return (profile and profile.role == "admin") or user.is_superuser